"""Gera uma planilha Excel de vendas para um mês escolhido, simulando cenários
reais de coleta de dados: completos ou com campos faltando, e padronizados ou
com erros de formatação (texto bagunçado, número em formato BR, datas soltas)
"""
from __future__ import annotations
import calendar
import argparse
import datetime as dt
import random
import sys
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Dados de referência
LOJAS = [
    ("L01", "Loja São Paulo", "São Paulo", "SP", "Sudeste"),
    ("L02", "Loja Rio de Janeiro", "Rio de Janeiro", "RJ", "Sudeste"),
    ("L03", "Loja Belo Horizonte", "Belo Horizonte", "MG", "Sudeste"),
    ("L04", "Loja Vitória", "Vitória", "ES", "Sudeste"),
    ("L05", "Loja Curitiba", "Curitiba", "PR", "Sul"),
    ("L06", "Loja Porto Alegre", "Porto Alegre", "RS", "Sul"),
    ("L07", "Loja Florianópolis", "Florianópolis", "SC", "Sul"),
    ("L08", "Loja Salvador", "Salvador", "BA", "Nordeste"),
    ("L09", "Loja Recife", "Recife", "PE", "Nordeste"),
    ("L10", "Loja Fortaleza", "Fortaleza", "CE", "Nordeste"),
]

PRODUTOS = [
    # produto_id, categoria, modelo, preco_unitario, custo_unitario, regiao_bias
    ("P01", "Notebook", "Lenovo", 3199.00, 2350.00, "Sudeste"),
    ("P02", "Notebook", "Positivo", 2199.00, 1550.00, "Nordeste"),
    ("P03", "Celular", "Samsung", 2499.00, 1800.00, "Sudeste"),
    ("P04", "Celular", "Motorola", 1799.00, 1250.00, "Nordeste"),
    ("P05", "Fone de Ouvido", "De Celular", 59.90, 28.00, "Sul"),
    ("P06", "Fone de Ouvido", "Headset", 249.90, 145.00, "Sudeste"),
    ("P07", "Carregador de Celular", "Tipo C", 44.90, 18.00, "Nordeste"),
    ("P08", "Controle de Videogame", "Xbox", 349.90, 215.00, "Sul"),
    ("P09", "Controle de Videogame", "PS4", 299.90, 175.00, "Sudeste"),
    ("P10", "Controle de Videogame", "PS5", 379.90, 235.00, "Sul"),
]

PRIMEIROS = ["Ana", "Bruno", "Carla", "Diego", "Elaine", "Fabio", "Gabriela", "Hugo",
             "Isabela", "João", "Karina", "Lucas", "Mariana", "Nelson", "Olívia",
             "Pedro", "Queila", "Rafael", "Sofia", "Thiago"]
SOBRENOMES = ["Silva", "Souza", "Oliveira", "Santos", "Pereira", "Costa", "Rodrigues",
              "Almeida", "Nascimento", "Lima", "Araújo", "Fernandes", "Carvalho"]

CAMPOS_VENDAS = ["venda_id", "data", "vendedor_id", "loja_id", "regiao", "produto_id",
                  "produto", "quantidade", "preco_unitario", "valor_total",
                  "custo_unitario", "custo_total"]

QTD_VENDEDORES = 100
QTD_VENDAS_MES = 450          # volume aproximado de vendas no mês escolhido
QTD_MINIMA_FALTANTES = 40
QTD_MAXIMA_FALTANTES = 60

def perguntar_opcao(mensagem: str, opcoes: dict[str, str]) -> str:
    """Mostra opções numeradas e repete a pergunta até receber uma resposta válida."""
    texto_opcoes = "\n".join(f"  {chave}) {desc}" for chave, desc in opcoes.items())
    while True:
        escolha = input(f"\n{mensagem}\n{texto_opcoes}\nEscolha: ").strip()
        if escolha in opcoes:
            return escolha
        print(f"Opção inválida: '{escolha}'. Digite uma das opções acima ({', '.join(opcoes)}).")

def perguntar_mes_ano() -> tuple[int, int]:
    """Pergunta mês e ano, validando com datetime/calendar antes de aceitar."""
    while True:
        bruto_mes = input("\nQual mês será analisado? (1 a 12): ").strip()
        bruto_ano = input("Qual ano? (ex: 2026): ").strip()
        try:
            mes, ano = int(bruto_mes), int(bruto_ano)
            # dt.date valida se mês/ano formam uma data real
            dt.date(ano, mes, 1)
            if ano < 2000 or ano > 2100:
                raise ValueError("ano fora de um intervalo razoável")
            return mes, ano
        except ValueError as erro:
            print(f"Entrada inválida ({erro}). Tente novamente com números, ex: mês=6, ano=2026.")

def coletar_respostas() -> dict:
    """Modo interativo (--interativo): pergunta modo/formatação/mês/ano no terminal."""
    print("=" * 60)
    print("GERADOR DE PLANILHA DE VENDAS PARA TESTE/ANÁLISE")
    print("=" * 60)

    modo = perguntar_opcao(
        "Modo de geração dos dados:",
        {"1": "Completo — todos os campos preenchidos (gastos completos)",
         "2": "Aleatório — parte dos registros pode ficar com campos faltando (~50)"},
    )
    formatacao = perguntar_opcao(
        "Formatação dos dados:",
        {"1": "Padronizados — textos e números já no formato correto",
         "2": "Com erro de formatação — texto bagunçado, número BR, datas soltas"},
    )
    mes, ano = perguntar_mes_ano()

    return {"modo": modo, "formatacao": formatacao, "mes": mes, "ano": ano}

def gerar_lojas() -> pd.DataFrame:
    """Cadastro fixo de lojas (dado de referência, não depende de sorteio)."""
    return pd.DataFrame(LOJAS, columns=["loja_id", "nome_loja", "cidade", "estado", "regiao"])

def gerar_vendedores(df_lojas: pd.DataFrame) -> pd.DataFrame:
    """Sorteia um nome para cada vendedor e distribui entre as lojas em
    round-robin (vendedor 1 -> loja 1, vendedor 2 -> loja 2, ...)."""
    linhas = []
    loja_regiao = dict(zip(df_lojas.loja_id, df_lojas.regiao))
    loja_nome = dict(zip(df_lojas.loja_id, df_lojas.nome_loja))
    for i in range(QTD_VENDEDORES):
        loja_id = df_lojas.loja_id.iloc[i % len(df_lojas)]
        nome = f"{random.choice(PRIMEIROS)} {random.choice(SOBRENOMES)}"
        linhas.append((f"V{i + 1:03d}", nome, loja_id, loja_nome[loja_id], loja_regiao[loja_id]))
    return pd.DataFrame(linhas, columns=["vendedor_id", "nome_vendedor", "loja_id", "nome_loja", "regiao"])

def gerar_produtos() -> pd.DataFrame:
    """Cadastro fixo de produtos, com margem calculada a partir do catálogo."""
    df = pd.DataFrame(PRODUTOS, columns=["produto_id", "categoria", "modelo",
                                          "preco_unitario", "custo_unitario", "regiao_bias"])
    df["margem_unitaria"] = (df.preco_unitario - df.custo_unitario).round(2)
    df["margem_%"] = (df.margem_unitaria / df.preco_unitario).round(4)
    return df.drop(columns="regiao_bias")

def gerar_vendas(mes: int, ano: int, df_vendedores: pd.DataFrame, df_produtos: pd.DataFrame) -> pd.DataFrame:
    """Sorteia QTD_VENDAS_MES vendas dentro do mês pedido: vendedor, produto,
    quantidade (1-3, com viés pra 1) e dia do mês, todos via random.*.
    A ordem dos sorteios por linha é: vendedor -> produto -> quantidade -> dia."""
    ultimo_dia = calendar.monthrange(ano, mes)[1]  # trata corretamente fevereiro/bissexto
    data_ini, data_fim = dt.date(ano, mes, 1), dt.date(ano, mes, ultimo_dia)
    dias_no_mes = (data_fim - data_ini).days

    produtos_info = df_produtos.set_index("produto_id")
    linhas = []
    for i in range(QTD_VENDAS_MES):
        vendedor = df_vendedores.sample(1).iloc[0]
        produto_id = random.choice(df_produtos.produto_id.tolist())
        info = produtos_info.loc[produto_id]
        qtd = random.choices([1, 2, 3], weights=[0.72, 0.22, 0.06], k=1)[0]
        data_venda = data_ini + dt.timedelta(days=random.randint(0, dias_no_mes))

        linhas.append({
            "venda_id": f"VD{i + 1:05d}",
            "data": data_venda,
            "vendedor_id": vendedor.vendedor_id,
            "loja_id": vendedor.loja_id,
            "regiao": vendedor.regiao,
            "produto_id": produto_id,
            "produto": f"{info.categoria} {info.modelo}",
            "quantidade": qtd,
            "preco_unitario": info.preco_unitario,
            "valor_total": round(qtd * info.preco_unitario, 2),
            "custo_unitario": info.custo_unitario,
            "custo_total": round(qtd * info.custo_unitario, 2),
        })

    df = pd.DataFrame(linhas, columns=CAMPOS_VENDAS)
    return df.sort_values("data").reset_index(drop=True)

def aplicar_dados_faltando(df: pd.DataFrame) -> pd.DataFrame:
    """Zera aleatoriamente um campo em ~40-60 registros, mantendo venda_id
    sempre presente (é a chave que identifica a linha na auditoria)."""
    resultado = df.copy()
    campos_opcionais = [c for c in resultado.columns if c != "venda_id"]
    n_faltantes = min(random.randint(QTD_MINIMA_FALTANTES, QTD_MAXIMA_FALTANTES), len(resultado))

    linhas_escolhidas = random.sample(range(len(resultado)), k=n_faltantes)
    for linha in linhas_escolhidas:
        campo = random.choice(campos_opcionais)
        resultado.at[linha, campo] = None

    print(f"  -> {n_faltantes} registros de Vendas ficaram com 1 campo faltando cada.")
    return resultado

def _bagunca_texto(valor):
    """Sorteia uma variante "suja" de um texto (maiúsculo, minúsculo,
    espaços nas pontas ou espaço duplo no meio)."""
    if pd.isna(valor):
        return valor
    texto = str(valor)
    variantes = [
        texto.upper(),
        texto.lower(),
        f"  {texto}  ",                      # espaços sobrando
        texto.replace(" ", "  "),             # espaço duplo no meio
    ]
    return random.choice(variantes)

def _bagunca_numero_brl(valor):
    """Formata um número no padrão BR como texto: 1.234,56 (em vez de 1234.56)."""
    if pd.isna(valor):
        return valor
    texto = f"{float(valor):,.2f}"                       # 1,234.56 (padrão US)
    texto = texto.replace(",", "§").replace(".", ",").replace("§", ".")  # -> 1.234,56
    return texto

def _bagunca_data(valor):
    """Sorteia um dos 4 formatos de data "soltos" (ver _FORMATOS_DATA_CONHECIDOS
    em qualidade.py — é esse o conjunto de formatos que o pipeline sabe reconhecer)."""
    if pd.isna(valor):
        return valor
    data = valor if isinstance(valor, dt.date) else pd.to_datetime(valor).date()
    formatos = ["%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"]
    return data.strftime(random.choice(formatos))


def _aplicar_erro_na_coluna(df: pd.DataFrame, coluna: str, colunas_texto: set[str], colunas_numero: set[str]) -> None:
    """Bagunça o conteúdo de UMA coluna, na mesma ordem/lógica da versão
    original — só extraído para fora do laço duplo por legibilidade, sem
    mudar quais nem quantas vezes random.* é chamado."""
    if coluna in colunas_texto:
        df[coluna] = df[coluna].map(_bagunca_texto)
    elif coluna in colunas_numero:
        # só bagunça uma parte das linhas, pra sobrar também número "limpo";
        # a coluna precisa virar 'object' antes, senão o pandas não deixa
        # misturar texto (ex: "1.234,56") dentro de uma coluna float64
        df[coluna] = df[coluna].astype(object)
        mascara = df[coluna].notna() & (pd.Series(range(len(df))) % 3 == 0)
        df.loc[mascara, coluna] = df.loc[mascara, coluna].map(_bagunca_numero_brl)
    elif coluna == "data":
        df[coluna] = df[coluna].map(_bagunca_data)


def aplicar_erros_formatacao(dfs: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    """Bagunça texto, número e data em cada tabela (nessa ordem: tabela por
    tabela, coluna por coluna, na mesma ordem de `dfs.items()`/`df.columns`)
    — a ordem importa porque com `random.seed()` fixo ela determina
    exatamente quais valores saem sorteados."""
    resultado = {nome: df.copy() for nome, df in dfs.items()}

    colunas_texto = {"nome_loja", "cidade", "estado", "regiao", "nome_vendedor",
                      "categoria", "modelo", "produto", "loja_id", "vendedor_id",
                      "produto_id", "venda_id"}
    colunas_numero = {"preco_unitario", "custo_unitario", "valor_total", "custo_total",
                       "margem_unitaria"}

    for df in resultado.values():
        for coluna in df.columns:
            _aplicar_erro_na_coluna(df, coluna, colunas_texto, colunas_numero)

    print("  -> Textos, números e datas bagunçados (maiúsc/minúsc, espaços, formato BR, datas soltas).")
    return resultado

def _checar_venda_id_duplicado(vendas: pd.DataFrame) -> None:
    """Imprime quantos venda_id duplicados existem (idealmente zero)."""
    duplicados = vendas["venda_id"].duplicated().sum()
    status = "(OK)" if duplicados == 0 else "(ATENÇÃO)"
    print(f"venda_id duplicado: {duplicados} {status}")


def _checar_campos_vazios(dfs: dict[str, pd.DataFrame]) -> None:
    """Imprime, para cada tabela, quantas células estão vazias."""
    for nome, df in dfs.items():
        faltando = int(df.isna().sum().sum())
        pct = round(faltando / df.size * 100, 2) if df.size else 0
        print(f"Campos vazios em {nome}: {faltando} ({pct}% das células)")


def _checar_modo_completo(vendas: pd.DataFrame, modo: str) -> None:
    """No modo "completo" (modo == "1"), nenhum campo deveria estar vazio;
    avisa se essa expectativa não se confirmar."""
    if modo != "1":
        return
    total_vazio = int(vendas.isna().sum().sum())
    if total_vazio == 0:
        status = "OK, nenhum campo vazio"
    else:
        status = f"ATENÇÃO: eram esperados 0 vazios, mas há {total_vazio}"
    print(f"Checagem modo completo: {status}")


def verificar_qualidade(dfs: dict[str, pd.DataFrame], modo: str) -> None:
    """Roda checagens básicas com pandas e imprime um resumo no terminal.
    Não interrompe a geração — é um relatório informativo, já que os modos
    'aleatório'/'com erro de formatação' geram problemas de propósito."""
    print("\n" + "-" * 60)
    print("VERIFICAÇÃO DE QUALIDADE (antes de salvar)")
    print("-" * 60)

    vendas = dfs["Vendas"]
    print(f"Linhas em Vendas: {len(vendas)}")

    _checar_venda_id_duplicado(vendas)
    _checar_campos_vazios(dfs)
    _checar_modo_completo(vendas, modo)

    print("-" * 60)

FONTE = "Arial"
FUNDO_CABECALHO = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FONTE_CABECALHO = Font(name=FONTE, bold=True, color="FFFFFF")
BORDA = Border(*(Side(style="thin", color="B7B7B7"),) * 4)


def _estilizar_cabecalho(ws) -> None:
    """Primeira linha em negrito, branco sobre fundo azul escuro."""
    for coluna in range(1, ws.max_column + 1):
        celula = ws.cell(row=1, column=coluna)
        celula.font = FONTE_CABECALHO
        celula.fill = FUNDO_CABECALHO


def _aplicar_bordas_e_fonte(ws) -> None:
    """Borda fina em toda célula; fonte padrão nas linhas de dado (a linha
    1 já ficou com a fonte do cabeçalho em _estilizar_cabecalho)."""
    for linha in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for celula in linha:
            celula.border = BORDA
            if celula.row > 1:
                celula.font = Font(name=FONTE)


def _ajustar_largura_colunas(ws) -> None:
    """Largura de cada coluna proporcional ao maior valor nela, entre 10 e 30."""
    for coluna in range(1, ws.max_column + 1):
        letra = get_column_letter(coluna)
        maior = max((len(str(c.value)) for c in ws[letra] if c.value is not None), default=10)
        ws.column_dimensions[letra].width = min(max(maior + 2, 10), 30)


def _congelar_cabecalho_e_habilitar_filtro(ws) -> None:
    """Congela a linha 1 (cabeçalho) e liga o autofiltro na planilha inteira."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _formatar_planilha(workbook, nomes_abas) -> None:
    """Aplica todo o estilo visual (cabeçalho, bordas, largura, filtro) em
    cada aba do workbook."""
    for nome in nomes_abas:
        ws = workbook[nome]
        _estilizar_cabecalho(ws)
        _aplicar_bordas_e_fonte(ws)
        _ajustar_largura_colunas(ws)
        _congelar_cabecalho_e_habilitar_filtro(ws)


def salvar_excel(dfs: dict[str, pd.DataFrame], caminho: Path) -> None:
    """Escreve cada DataFrame como uma aba e aplica formatação visual.

    Duas passadas: primeiro o pandas escreve os dados (rápido, sem estilo),
    depois o openpyxl reabre o arquivo só para aplicar o estilo — é mais
    simples que formatar célula a célula durante a escrita.
    """
    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        for nome, df in dfs.items():
            df.to_excel(writer, sheet_name=nome, index=False)

    workbook = load_workbook(caminho)
    _formatar_planilha(workbook, dfs.keys())
    workbook.save(caminho)

def gerar_modelo(mes: int, ano: int, modo: str, formatacao: str, pasta_destino: Path) -> Path:
    """Gera um único cenário; usado tanto pela CLI quanto pelo GitHub Actions."""

    print("\nGerando dados de referência (Lojas, Vendedores, Produtos)...")
    df_lojas = gerar_lojas()
    df_vendedores = gerar_vendedores(df_lojas)
    df_produtos = gerar_produtos()

    print(f"Gerando vendas para {mes:02d}/{ano}...")
    df_vendas = gerar_vendas(mes, ano, df_vendedores, df_produtos)

    if modo == "2":
        print("Aplicando modo ALEATÓRIO (campos faltando)...")
        df_vendas = aplicar_dados_faltando(df_vendas)

    dfs = {"Lojas": df_lojas, "Vendedores": df_vendedores, "Produtos": df_produtos, "Vendas": df_vendas}

    if formatacao == "2":
        print("Aplicando formatação COM ERRO (texto/número/data bagunçados)...")
        dfs = aplicar_erros_formatacao(dfs)
    else:
        print("Mantendo formatação PADRONIZADA.")

    verificar_qualidade(dfs, modo)

    sufixo_modo = "completo" if modo == "1" else "aleatorio"
    sufixo_formato = "padronizado" if formatacao == "1" else "com_erro"
    nome_arquivo = f"vendas_{ano}{mes:02d}_{sufixo_modo}_{sufixo_formato}.xlsx"

    pasta_destino.mkdir(parents=True, exist_ok=True)
    caminho_saida = pasta_destino / nome_arquivo

    salvar_excel(dfs, caminho_saida)
    print(f"Planilha gerada: {caminho_saida.resolve()}")
    return caminho_saida


def _semear(seed: int) -> None:
    """Semeia os DOIS geradores de números aleatórios usados no script.

    `random.seed()` sozinho NÃO é suficiente: `df.sample()` (usado em
    gerar_vendas() para sortear o vendedor de cada venda) usa o RNG
    global do numpy, que é totalmente independente do `random` da
    biblioteca padrão. Sem semear os dois, só parte dos sorteios (produto,
    quantidade, dia do mês) é reproduzível — o vendedor sorteado muda a
    cada execução, mesmo com a mesma seed.
    """
    random.seed(seed)
    np.random.seed(seed)


def main() -> None:
    """CLI: modo padrão gera os 2 cenários de CI; --interativo pergunta no terminal."""
    parser = argparse.ArgumentParser(description="Gera planilhas de teste para a esteira de dados.")
    parser.add_argument("--mes", type=int, default=6, choices=range(1, 13), help="Mês de referência (padrão: 6).")
    parser.add_argument("--ano", type=int, default=2026, help="Ano de referência (padrão: 2026).")
    parser.add_argument("--destino", type=Path, default=Path(__file__).resolve().parent.parent / "data" / "raw", help="Diretório de saída.")
    parser.add_argument("--seed", type=int, default=20260727, help="Semente para resultados reproduzíveis.")
    parser.add_argument("--interativo", action="store_true", help="Mantém o modo antigo de perguntas no terminal.")
    args = parser.parse_args()

    if args.interativo:
        respostas = coletar_respostas()
        _semear(args.seed)
        gerar_modelo(respostas["mes"], respostas["ano"], respostas["modo"], respostas["formatacao"], args.destino)
        return

    # Modo padrão, próprio para CI: sempre produz os dois cenários solicitados.
    print("Gerando os dois modelos não interativos para validação automática...")
    _semear(args.seed)
    gerar_modelo(args.mes, args.ano, "1", "1", args.destino)
    _semear(args.seed + 1)
    gerar_modelo(args.mes, args.ano, "2", "2", args.destino)
    print("Concluído: modelo íntegro e modelo com pendências/formatação incorreta.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nOperação cancelada pelo usuário.")
        sys.exit(1)
